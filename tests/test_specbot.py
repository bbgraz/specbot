"""Functional test suite for SpecBot — runs offline (no OpenAI key needed).

Usage:  python tests/test_specbot.py
Extra dev dependency for the SMTP round-trip test:  pip install aiosmtpd
"""
from __future__ import annotations

import copy
import json
import os
import sys
import tempfile
import time
import traceback
from pathlib import Path

APP_DIR = Path(__file__).resolve().parent.parent
SCRATCH = Path(tempfile.mkdtemp(prefix="specbot_test_"))
sys.path.insert(0, str(APP_DIR))

# Isolate side effects from the repo
os.environ["SPECBOT_EXPORT_DIR"] = str(SCRATCH / "exports")
os.environ["SPECBOT_WIP_PATH"] = str(SCRATCH / "wip_records.json")
os.environ["SPECBOT_TECHPACK_DIR"] = str(SCRATCH / "tech_packs")
os.environ.pop("OPENAI_API_KEY", None)
os.environ.pop("RESEND_API_KEY", None)
os.environ.pop("SMTP_HOST", None)
os.environ.pop("TEST_EMAIL_RECIPIENT", None)

RESULTS: list[tuple[str, bool, str]] = []


def check(name: str, fn):
    try:
        fn()
        RESULTS.append((name, True, ""))
    except Exception as exc:  # noqa: BLE001
        RESULTS.append((name, False, f"{exc.__class__.__name__}: {exc}\n{traceback.format_exc(limit=3)}"))


SAMPLE_TP = {
    "style_number": "TST-001",
    "style_name": "Demo Tee",
    "garment_type": "crewneck tee",
    "fabric": "180gsm cotton jersey",
    "sample_size": "M",
    "garment_summary": "Short-sleeve crewneck tee.",
    "measurements": [
        {"pom": "Chest Width", "description": "1in below armhole", "target": "21", "tolerance_plus": "0.25", "tolerance_minus": "0.25", "source": "derived_from_input", "notes": ""},
        {"pom": "Armhole Depth", "description": "straight", "target": "9.5", "tolerance_plus": "0.25", "tolerance_minus": "0.25", "source": "inferred_from_standard_practice", "notes": ""},
        {"pom": "Sleeve Length", "description": "from shoulder", "target": "8.25", "tolerance_plus": "0.25", "tolerance_minus": "0.25", "source": "placeholder_for_review", "notes": ""},
    ],
    "construction_notes": [{"note": "Coverstitch hem", "source": "inferred_from_standard_practice"}],
    "bom": [{"component": "Self fabric", "material": "cotton jersey", "placement": "body", "notes": "", "source": "derived_from_input"}],
    "change_log": [],
    "assumptions": ["Assumed standard fit"],
    "missing_information": ["Neck rib height unclear"],
}


# ---------------------------------------------------------------- imports
def t_imports():
    import gpt_service, excel_exporter, fit_update_service, wip_store, email_sender, mock_data  # noqa
check("All modules import cleanly", t_imports)


# ---------------------------------------------------------------- gpt_service guards
def t_gpt_no_key():
    from gpt_service import analyze_sketch
    try:
        analyze_sketch(None, {"style_name": "x"})
    except RuntimeError as exc:
        assert "OPENAI_API_KEY" in str(exc)
        return
    raise AssertionError("analyze_sketch should raise RuntimeError without API key")
check("gpt_service refuses cleanly without API key", t_gpt_no_key)

def t_safe_json():
    from gpt_service import _safe_json_loads
    assert _safe_json_loads('```json\n{"a": 1}\n```') == {"a": 1}
    assert _safe_json_loads('noise {"a": 1} trailing') == {"a": 1}
check("gpt_service tolerates fenced/dirty JSON", t_safe_json)


# ---------------------------------------------------------------- spec blocks
def t_spec_match():
    from spec_blocks import match_category
    assert match_category("crewneck tee") == "tee"
    assert match_category("Zip Hoodie") == "sweatshirt"
    assert match_category("oxford button-down shirt") == "woven_shirt"
    assert match_category("slim chino pant") == "pants"
    assert match_category("mystery garment") == "default"
check("Spec blocks: garment type maps to category (longest keyword wins)", t_spec_match)

def t_spec_size_projection():
    from spec_blocks import get_spec_block
    m_block = get_spec_block("tee", "M")
    l_block = get_spec_block("tee", "L")
    chest_m = next(m["target"] for m in m_block["measurements"] if m["pom"] == "Chest Width")
    chest_l = next(m["target"] for m in l_block["measurements"] if m["pom"] == "Chest Width")
    assert float(chest_l) - float(chest_m) == 1.0  # chest grade rule = 1.0/size
    assert all(m["source"] == "inferred_from_standard_practice" for m in m_block["measurements"])
check("Spec blocks: targets project across sizes via grade rules", t_spec_size_projection)

def t_grounding():
    from spec_blocks import ground_measurements, get_spec_block
    block = get_spec_block("tee", "M")
    n_block = len(block["measurements"])
    ai = [
        # plausible adjustment (within 35% of 21) -> accepted
        {"pom": "Chest Width", "target": "22", "source": "derived_from_input", "notes": "wide fit per sketch"},
        # implausible (9.5 -> 30) -> rejected, standard kept, flagged
        {"pom": "Armhole Depth", "target": "30", "source": "derived_from_input"},
        # extra POM not in block, not derived -> downgraded to placeholder
        {"pom": "Pocket Width", "target": "5", "source": "inferred_from_standard_practice"},
    ]
    grounded, notes = ground_measurements(ai, "tee", "M")
    by_pom = {m["pom"]: m for m in grounded}
    assert by_pom["Chest Width"]["target"] == "22"
    assert by_pom["Armhole Depth"]["target"] == "9.5"
    assert by_pom["Armhole Depth"]["source"] == "placeholder_for_review"
    assert any("Armhole Depth" in n for n in notes)
    assert by_pom["Pocket Width"]["source"] == "placeholder_for_review"
    assert len(grounded) == n_block + 1  # every block POM present + 1 extra
check("Grounding: plausible AI values kept, implausible rejected, extras flagged", t_grounding)

def t_offline_draft():
    from spec_blocks import build_offline_draft
    draft = build_offline_draft({"garment_type": "hoodie", "sample_size": "L", "fabric": "fleece"})
    assert draft["suggested_measurements"] and draft["bom_items"] and draft["construction_notes"]
    assert any("offline" in a.lower() or "no ai" in a.lower() for a in draft["assumptions"])
    assert draft["missing_information"]
    poms = [m["pom"] for m in draft["suggested_measurements"]]
    assert "Hood Height" in poms
check("Offline draft: full analysis shape from spec block, honest caveats", t_offline_draft)


# ---------------------------------------------------------------- brand library + persistence
def t_brand_library():
    from brand_library import build_grounding, grounding_for_prompt, grounding_report
    g = build_grounding("crewneck tee", "cotton jersey")
    prompt = grounding_for_prompt(g)
    assert prompt and "FBR-" in prompt  # cites real brand-library codes
    report = grounding_report(g)
    assert isinstance(report, dict) and report
check("Brand library: grounding block + report built from mock brand data", t_brand_library)

def t_tech_pack_store():
    import tech_pack_store
    tech_pack_store.TECH_PACK_DIR.mkdir(parents=True, exist_ok=True)
    saved = tech_pack_store.save_tech_pack(copy.deepcopy(SAMPLE_TP))
    assert Path(saved).is_file()
    loaded = tech_pack_store.load_tech_pack("TST-001")
    assert loaded and loaded["style_name"] == "Demo Tee"
    listing = tech_pack_store.list_tech_packs()
    assert any(r.get("style_number") == "TST-001" for r in listing)
    assert tech_pack_store.delete_tech_pack("TST-001")
    assert tech_pack_store.load_tech_pack("TST-001") is None
check("Tech pack store: save, load, list, delete round-trip", t_tech_pack_store)


# ---------------------------------------------------------------- fit update (rule-based)
def t_fit_delta():
    from fit_update_service import apply_fitting_notes
    original = copy.deepcopy(SAMPLE_TP)
    revised = apply_fitting_notes(original, "Raise armhole by 0.5. Chest width -0.25.")
    assert original["measurements"][1]["target"] == "9.5", "original mutated!"
    m = {r["pom"]: r["target"] for r in revised["measurements"]}
    assert m["Armhole Depth"] == "10", m
    assert m["Chest Width"] == "20.75", m
    assert len(revised["change_log"]) == 2
    assert all(e["old_value"] and e["new_value"] for e in revised["change_log"])
check("Fitting notes: deltas applied + change log, original untouched", t_fit_delta)

def t_fit_set_and_new():
    from fit_update_service import apply_fitting_notes
    revised = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), "set sleeve length to 9\nfront neck drop +0.5")
    m = {r["pom"].lower(): r for r in revised["measurements"]}
    assert m["sleeve length"]["target"] == "9"
    new_row = next(r for r in revised["measurements"] if "neck" in r["pom"].lower())
    assert new_row["source"] == "fitting_note"
    assert len(revised["change_log"]) == 2
check("Fitting notes: absolute set + unknown POM added as fitting_note row", t_fit_set_and_new)

def t_fit_unparseable():
    from fit_update_service import apply_fitting_notes
    revised = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), "Fit looked great overall, model liked it")
    assert len(revised["change_log"]) == 1
    assert "no actionable update" in revised["change_log"][0]["reason"]
check("Fitting notes: unparseable note still logged (audit trail)", t_fit_unparseable)

def t_fit_empty():
    from fit_update_service import apply_fitting_notes
    revised = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), "   ")
    assert revised == SAMPLE_TP
check("Fitting notes: empty input is a no-op", t_fit_empty)

def t_fit_fractions():
    from fit_update_service import apply_fitting_notes, _parse_amount
    assert _parse_amount("3/8") == 0.375
    assert _parse_amount("1 1/2") == 1.5
    assert _parse_amount("½") == 0.5
    assert _parse_amount('1/4"') == 0.25
    notes = 'let out chest 1/2\nRaise armhole depth 1/4"\nshorten sleeve length by 3/8'
    revised = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), notes)
    m = {r["pom"]: r["target"] for r in revised["measurements"]}
    assert m["Chest Width"] == "21.5", m
    assert m["Armhole Depth"] == "9.75", m
    assert m["Sleeve Length"] == "7.875", m   # 8.25 - 0.375; the old parser read 3/8 as 3
check("Fitting notes: fractions parse correctly (1/2, 1/4\", 3/8, no 'by' needed)", t_fit_fractions)

def t_fit_plausibility_cap():
    from fit_update_service import apply_fitting_notes
    revised = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), "shorten sleeve length by 5")
    m = {r["pom"]: r["target"] for r in revised["measurements"]}
    assert m["Sleeve Length"] == "8.25", "implausible delta must not be applied"
    assert any("NOT APPLIED" in e["reason"] for e in revised["change_log"])
check("Fitting notes: implausible delta blocked, flagged NOT APPLIED", t_fit_plausibility_cap)

def t_fit_no_silent_drops():
    from fit_update_service import apply_fitting_notes
    notes = "raise armhole depth by 0.25\nmove grainline 3 degrees on sleeve panel"
    revised = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), notes)
    reasons = " | ".join(e["reason"] for e in revised["change_log"])
    assert "grainline" in reasons and "could not parse" in reasons, reasons
    assert len(revised["change_log"]) == 2  # one applied + one unparsed, zero silent
check("Fitting notes: unparseable actionable lines logged, never dropped", t_fit_no_silent_drops)


# ---------------------------------------------------------------- excel export
def t_excel():
    from excel_exporter import export_tech_pack_to_excel
    from fit_update_service import apply_fitting_notes
    tp = apply_fitting_notes(copy.deepcopy(SAMPLE_TP), "raise armhole by 0.5")
    path = export_tech_pack_to_excel(tp)
    assert Path(path).is_file() and Path(path).stat().st_size > 5000
    from openpyxl import load_workbook
    wb = load_workbook(path)
    expected = ["Cover", "Measurements", "BOM", "Construction", "Grading", "Annotations", "Change Log", "Assumptions and Missing"]
    assert wb.sheetnames == expected, wb.sheetnames
    for name in expected:
        ws = wb[name]
        assert "TST-001" in str(ws.cell(row=2, column=1).value), f"style header missing on {name}"
        assert ws.freeze_panes, f"no frozen panes on {name}"
    meas = wb["Measurements"]
    grid = [[c.value for c in row] for row in meas.iter_rows(min_row=5, max_row=8)]
    assert grid[0][:3] == ["POM", "Description", "Target"]
    targets = {r[0]: r[2] for r in grid[1:]}
    assert targets["Armhole Depth"] == "10"
    cl = wb["Change Log"]
    cl_rows = [[c.value for c in row] for row in cl.iter_rows(min_row=6, max_row=6)]
    assert cl_rows[0][4] == "9.5" and cl_rows[0][5] == "10"
    globals()["_EXPORT_PATH"] = path
check("Excel export: 6 sheets, headers, frozen rows, revised values, change log", t_excel)


def t_excel_sketch_and_rev():
    from excel_exporter import export_tech_pack_to_excel
    from PIL import Image
    import io
    img = Image.new("RGB", (600, 400), "white")
    buf = io.BytesIO(); img.save(buf, format="JPEG")
    tp = copy.deepcopy(SAMPLE_TP); tp["rev"] = 2
    path = export_tech_pack_to_excel(tp, sketch_bytes=buf.getvalue())
    assert "_rev2_" in Path(path).name
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert len(wb["Cover"]._images) == 1, "sketch must be embedded on Cover"
    bom = wb["BOM"]
    headers = [c.value for c in bom[5]]
    assert "Qty/Consumption" in headers and "Supplier / Article #" in headers
check("Excel: sketch embedded on Cover, rev in filename, full BOM columns", t_excel_sketch_and_rev)

def t_grading_sanity():
    from mock_data import build_graded_table, NUMERIC_SIZE_RUN
    meas = [
        {"pom": "Neck Rib Height", "target": "0.75"},
        {"pom": "Chest Width", "target": "21"},
        {"pom": "Waist Width", "target": "16.5"},
    ]
    rows = build_graded_table(meas, sample_size="M")
    rib = next(r for r in rows if r["POM"] == "Neck Rib Height")
    assert rib["XS"] == "0.75" and rib["XXL"] == "0.75", "trim heights must not grade"
    assert all(float(v) >= 0 for r in rows for k, v in r.items() if k not in ("POM", "Grade rule (in)"))
    num = build_graded_table(meas, sample_size="32", size_run=NUMERIC_SIZE_RUN)
    waist = next(r for r in num if r["POM"] == "Waist Width")
    assert waist["32"] == "16.5" and waist["34"] == "17.5", waist
check("Grading: zero-grade trims, floor at 0, numeric run for bottoms", t_grading_sanity)


# ---------------------------------------------------------------- wip store
def t_wip():
    import wip_store
    wip_store.WIP_PATH.unlink(missing_ok=True)
    assert wip_store.load_wip_records() == []
    rec = {"style_number": "TST-001", "style_name": "Demo Tee", "factory_name": "Lotus Apparel Manufacturing",
           "contact_name": "Anong S.", "status": "Draft", "tech_pack_file": "x.xlsx"}
    records = wip_store.add_or_update_wip_record(rec)
    assert len(records) == 1 and records[0]["last_update"]
    records = wip_store.add_or_update_wip_record({"style_number": "TST-001", "status": "Sent to Factory"})
    assert len(records) == 1 and records[0]["status"] == "Sent to Factory"
    assert records[0]["style_name"] == "Demo Tee"  # merge preserved other fields
    on_disk = json.loads(wip_store.WIP_PATH.read_text())
    assert on_disk[0]["status"] == "Sent to Factory"
    try:
        wip_store.add_or_update_wip_record({"style_name": "no number"})
        raise AssertionError("should reject missing style_number")
    except ValueError:
        pass
check("WIP store: add, upsert-merge, persist to JSON, reject bad record", t_wip)


# ---------------------------------------------------------------- factory contacts
def t_contacts():
    data = json.loads((APP_DIR / "factory_contacts.json").read_text())
    assert len(data) == 5
    for f in data:
        assert all(k in f for k in ("factory_id", "factory_name", "country", "specialty", "contacts"))
        assert len(f["contacts"]) == 2
        for c in f["contacts"]:
            assert all(k in c for k in ("name", "title", "email"))
            assert c["email"].endswith(".example.com") or "example" in c["email"], f"non-fictional email? {c['email']}"
check("Factory contacts: 5 factories x 2 contacts, fictional emails", t_contacts)


# ---------------------------------------------------------------- email sender
def t_email_guards():
    from email_sender import send_factory_email
    r = send_factory_email("real.factory@example.com", "s", "b")
    assert not r["ok"] and "TEST_EMAIL_RECIPIENT" in r["error"]
    os.environ["TEST_EMAIL_RECIPIENT"] = "owner@test.example.com"
    r = send_factory_email("real.factory@example.com", "s", "b")
    assert not r["ok"] and "transport" in r["error"].lower()
    os.environ.pop("TEST_EMAIL_RECIPIENT")
check("Email: refuses without TEST_EMAIL_RECIPIENT / without transport", t_email_guards)

def t_email_smtp_roundtrip():
    from aiosmtpd.controller import Controller

    received = []

    class Sink:
        async def handle_DATA(self, server, session, envelope):
            received.append(envelope)
            return "250 OK"

    controller = Controller(Sink(), hostname="127.0.0.1", port=8825)
    controller.start()
    try:
        os.environ["TEST_EMAIL_RECIPIENT"] = "owner@test.example.com"
        os.environ["SMTP_HOST"] = "127.0.0.1"
        os.environ["SMTP_PORT"] = "8825"
        os.environ["EMAIL_FROM"] = "specbot@test.example.com"
        from email_sender import send_factory_email
        r = send_factory_email(
            "anong.s@lotus-apparel.example.com", "Tech pack TST-001", "Please review.",
            attachment_path=globals().get("_EXPORT_PATH"),
        )
        assert r["ok"], r
        assert r["to"] == "owner@test.example.com"  # forced routing
        time.sleep(0.3)
        assert len(received) == 1
        env = received[0]
        assert env.rcpt_tos == ["owner@test.example.com"], env.rcpt_tos
        raw = env.content.decode("utf-8", "replace")
        assert "anong.s@lotus-apparel.example.com" in raw  # intended recipient in body
        assert "test-mode" in raw
        assert "techpack_TST-001" in raw  # attachment present
    finally:
        controller.stop()
        for k in ("SMTP_HOST", "SMTP_PORT", "EMAIL_FROM", "TEST_EMAIL_RECIPIENT"):
            os.environ.pop(k, None)
check("Email: SMTP round-trip — forced to test recipient, banner + Excel attached", t_email_smtp_roundtrip)


# ---------------------------------------------------------------- streamlit UI smoke
def t_ui_boot():
    from streamlit.testing.v1 import AppTest
    os.chdir(APP_DIR)
    at = AppTest.from_file(str(APP_DIR / "app.py"), default_timeout=60)
    at.run()
    assert len(at.exception) == 0, list(at.exception)
    body = " ".join(str(m.value) for m in at.markdown)
    assert "SpecBot" in body
    globals()["_AT"] = at
check("Streamlit app boots headless with no exception", t_ui_boot)

def t_ui_form_validation():
    at = globals()["_AT"]
    # submit the style form empty -> should show required-fields error
    buttons = [b for b in at.button if "Generate" in (b.label or "")]
    assert buttons, "Generate Tech Pack button not found"
    buttons[0].click().run()
    assert len(at.exception) == 0
    errs = " ".join(str(e.value) for e in at.error)
    assert "required" in errs.lower(), f"expected validation error, got: {errs!r}"
check("UI: empty Generate submit shows validation error (no crash)", t_ui_form_validation)

def t_ui_no_key_generate():
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(APP_DIR / "app.py"), default_timeout=60)
    at.run()
    at.text_input(key="form_style_name").set_value("Demo Tee")
    at.text_input(key="form_style_number").set_value("TST-001")
    at.text_input(key="form_garment_type").set_value("crewneck tee")
    subs = [b for b in at.button if "Generate" in (b.label or "")]
    subs[0].click().run()
    assert len(at.exception) == 0
    warns = " ".join(str(w.value) for w in at.warning)
    assert "offline draft" in warns.lower(), warns
    tp = at.session_state["tech_pack"]
    assert tp["style_number"] == "TST-001"
    assert len(tp["measurements"]) >= 5
    assert all(m["source"] == "inferred_from_standard_practice" for m in tp["measurements"])
check("UI: Generate without API key falls back to offline spec-block draft", t_ui_no_key_generate)

def t_ui_load_demo():
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(APP_DIR / "app.py"), default_timeout=60)
    at.run()
    demo = [b for b in at.button if "Demo" in (b.label or "")]
    assert demo, "Load Demo Tech Pack button not found"
    demo[0].click().run()
    assert len(at.exception) == 0
    tp = at.session_state["tech_pack"]
    assert tp["style_number"] == "TST-001" and tp["garment_type"] == "crewneck tee"
    assert len(tp["measurements"]) >= 10 and tp["bom"] and tp["construction_notes"]
    assert tp["assumptions"] and tp["missing_information"]
check("UI: Load Demo Tech Pack populates a full tech pack offline", t_ui_load_demo)


def t_description_in_prompt():
    from gpt_service import _build_user_content
    meta = {"style_name": "Boxy Tee", "garment_type": "crewneck tee",
            "style_description": "Oversized boxy fit, dropped shoulders, raw-edge hems"}
    content = _build_user_content(None, None, meta)
    text = content[0]["text"]
    assert "Design intent stated by the designer" in text
    assert "dropped shoulders" in text
    assert "derived_from_input" in text
    # and absent when not provided
    text2 = _build_user_content(None, None, {"style_name": "X", "garment_type": "tee"})[0]["text"]
    assert "Design intent stated" not in text2
check("Style description feeds GPT prompt as designer-stated fact", t_description_in_prompt)

def t_description_offline_and_ui():
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(APP_DIR / "app.py"), default_timeout=60)
    at.run()
    at.text_input(key="form_style_name").set_value("Boxy Tee")
    at.text_input(key="form_style_number").set_value("SS26-014")
    at.text_input(key="form_garment_type").set_value("crewneck tee")
    at.text_area(key="form_style_description").set_value("Oversized boxy fit, raw-edge hems")
    subs = [b for b in at.button if "Generate" in (b.label or "")]
    subs[0].click().run()
    assert len(at.exception) == 0
    tp = at.session_state["tech_pack"]
    assert tp["style_description"] == "Oversized boxy fit, raw-edge hems"
    assert any("NOT interpreted" in m for m in tp["missing_information"]), \
        "offline draft must flag that the description was not analyzed"
check("UI: description captured on tech pack; offline draft flags it honestly", t_description_offline_and_ui)

def t_ui_workflow_navigation():
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(APP_DIR / "app.py"), default_timeout=60)
    at.run()
    assert at.session_state["nav_stage"] == "intake", "app starts at Style Intake"
    # gated stages redirect to intake when no style is loaded
    for stage in ("techpack", "fit", "send"):
        at.radio(key="nav_stage").set_value(stage).run()
        assert len(at.exception) == 0
        assert any("Style Intake" in str(i.value) for i in at.info), f"{stage} should be gated"
    # ungated stages render without a style
    for stage in ("wip", "library"):
        at.radio(key="nav_stage").set_value(stage).run()
        assert len(at.exception) == 0
check("UI: workflow nav — starts at Intake, gates stages, all stages render", t_ui_workflow_navigation)

def t_ui_preview_apply_flow():
    from streamlit.testing.v1 import AppTest
    at = AppTest.from_file(str(APP_DIR / "app.py"), default_timeout=60)
    at.run()
    demo = [b for b in at.button if "Demo" in (b.label or "")]
    demo[0].click().run()
    assert at.session_state["nav_stage"] == "techpack", "demo load should advance to Tech Pack"
    at.radio(key="nav_stage").set_value("fit").run()
    at.text_area(key="fitting_notes_input").set_value("raise armhole depth by 1/4\nshorten body length by 40")
    preview = [b for b in at.button if b.key == "preview_fitting_btn"]
    assert preview, "Preview Changes button not found"
    preview[0].click().run()
    assert len(at.exception) == 0
    # nothing applied yet
    tp = at.session_state["tech_pack"]
    armhole = next(m for m in tp["measurements"] if m["pom"] == "Armhole Depth")
    assert armhole["target"] == "9.5", "preview must not mutate the tech pack"
    apply_btn = [b for b in at.button if b.key == "apply_fitting_btn"]
    assert apply_btn, "Apply button should appear after preview"
    apply_btn[0].click().run()
    assert len(at.exception) == 0
    tp = at.session_state["tech_pack"]
    armhole = next(m for m in tp["measurements"] if m["pom"] == "Armhole Depth")
    assert armhole["target"] == "9.75"
    body = next(m for m in tp["measurements"] if m["pom"] == "Body Length from HPS")
    assert body["target"] == "28.5", "implausible 40in change must be blocked"
    assert tp["rev"] == 1 and tp["measurement_history"], "rev + snapshot recorded"
    assert all(e.get("stage") for e in tp["change_log"]), "stage stamped on entries"
check("UI: preview shows diff without mutating; apply commits rev + snapshot", t_ui_preview_apply_flow)


# ---------------------------------------------------------------- report
print("\n" + "=" * 72)
passed = sum(1 for _, ok, _ in RESULTS if ok)
for name, ok, err in RESULTS:
    print(f"{'PASS' if ok else 'FAIL'}  {name}")
    if not ok:
        print("      " + err.replace("\n", "\n      "))
print("=" * 72)
print(f"{passed}/{len(RESULTS)} passed")
sys.exit(0 if passed == len(RESULTS) else 1)
